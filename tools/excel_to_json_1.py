from pathlib import Path
import json
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent

EXCEL_FILE = PROJECT_ROOT / "database" / "excel" / "products_master.xlsx"
OUTPUT_JSON = PROJECT_ROOT / "mobile" / "assets" / "data" / "products.json"

SHEET_NAME = "Products"


def bool_value(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in (
        "true",
        "yes",
        "1",
        "да",
        "x",
    )


def clean(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if value == "":
            return None

    return value


def to_float(value):
    value = clean(value)

    if value is None:
        return None

    try:
        return float(str(value).replace(",", "."))
    except:
        return None


def split_list(value):
    value = clean(value)

    if value is None:
        return []

    return [
        x.strip()
        for x in str(value).split(";")
        if x.strip()
    ]


def build_image_path(category, subcategory, brand, variant, product_id):
    assets = PROJECT_ROOT / "mobile" / "assets"

    if category == "beer" and brand and variant:
        rel = Path("beer") / brand / f"{variant}.jpg"
        return (
            f"assets/{rel.as_posix()}"
            if (assets / rel).exists()
            else None
        )

    if category == "food" and subcategory and product_id:
        rel = Path("food") / subcategory / f"{product_id}.jpg"
        return (
            f"assets/{rel.as_posix()}"
            if (assets / rel).exists()
            else None
        )

    return None


def load_products():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    try:
        ws = wb[SHEET_NAME]
    except KeyError:
        match = next(
            (name for name in wb.sheetnames if name.lower() == SHEET_NAME.lower()),
            None,
        )
        if match is None:
            raise KeyError(
                f"Не намирам лист '{SHEET_NAME}' в {EXCEL_FILE.name}. "
                f"Налични листове: {wb.sheetnames}"
            )
        ws = wb[match]

    headers = {}

    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value

        if h:
            headers[str(h).strip()] = col

    products = []

    for row in range(2, ws.max_row + 1):

        product_id = clean(ws.cell(row, headers["id"]).value)

        if not product_id:
            continue

        category = clean(
            ws.cell(row, headers["category"]).value
        )

        subcategory = clean(
            ws.cell(row, headers["subcategory"]).value
        )

        if category:
            category = category.lower()

        if subcategory:
            subcategory = subcategory.lower()

        brand = clean(ws.cell(row, headers["brand"]).value)
        variant = clean(ws.cell(row, headers["variant"]).value)

        product = {
            "id": product_id,

            "category": category,
            "subcategory": subcategory,

            "brand": brand,
            "variant": variant,

            "nameBg": clean(
                ws.cell(row, headers["nameBg"]).value
            ),

            "nameEn": clean(
                ws.cell(row, headers["nameEn"]).value
            ),

            "descriptionBg": clean(
                ws.cell(row, headers["descriptionBg"]).value
            ),

            "descriptionEn": clean(
                ws.cell(row, headers["descriptionEn"]).value
            ),

            "priceEur": to_float(
                ws.cell(row, headers["priceEur"]).value
            ),

            "quantity": to_float(
                ws.cell(row, headers["quantity"]).value
            ),

            "unit": clean(
                ws.cell(row, headers["unit"]).value
            ),

            "image": build_image_path(
                category,
                subcategory,
                brand,
                variant,
                product_id,
            ),

            "featured": bool_value(
                ws.cell(row, headers["featured"]).value
            ),

            "isNew": bool_value(
                ws.cell(row, headers["isNew"]).value
            ),

            "isRecommended": bool_value(
                ws.cell(row, headers["isRecommended"]).value
            ),

            "available": bool_value(
                ws.cell(row, headers["available"]).value
            ),

            "tags": split_list(
                ws.cell(row, headers["tags"]).value
            )
            if "tags" in headers
            else [],

            "allergens": split_list(
                ws.cell(row, headers["allergens"]).value
            )
            if "allergens" in headers
            else [],
        }

        products.append(product)

    return products


def main():
    products = load_products()

    OUTPUT_JSON.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with open(
        OUTPUT_JSON,
        "w",
        encoding="utf-8",
    ) as f:
        json.dump(
            products,
            f,
            ensure_ascii=False,
            indent=2,
        )

    print("=" * 50)
    print(f"Generated {len(products)} products")
    print(f"Output: {OUTPUT_JSON}")
    print("=" * 50)


if __name__ == "__main__":
    main()